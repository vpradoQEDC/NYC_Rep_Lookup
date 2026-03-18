import sys, os, re, time, requests, pandas as pd
from bs4 import BeautifulSoup
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QLabel, QFileDialog, QProgressBar, QTableWidget,
    QTableWidgetItem, QHeaderView, QMessageBox, QFrame, QSizePolicy
)
from PyQt5.QtCore import Qt, QThread, pyqtSignal
from PyQt5.QtGui import QPixmap, QIcon, QColor, QFont
import openpyxl
from openpyxl.styles import Font as XLFont, PatternFill, Alignment as XLAlign, Border, Side
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────────────────────
#  Helpers
# ─────────────────────────────────────────────────────────────
def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except AttributeError:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)


def detect_col(columns, *keywords):
    for col in columns:
        lower = col.lower()
        if any(k in lower for k in keywords):
            return col
    return None


def is_valid_address(address):
    if not address:
        return False
    bad = {"-", "'", "'-", "'\\-", "nan", "n/a", "none", "unknown", "tbd", "", "0"}
    a = re.sub(r"^['\-\s]+", "", address).strip().lower()
    if a in bad or len(a) < 5:
        return False
    if not re.search(r'\d', a):
        return False
    return True


# ─────────────────────────────────────────────────────────────
#  ZIP Lookup — stores ALL combos per ZIP for shared-district logic
# ─────────────────────────────────────────────────────────────
def load_zip_lookup():
    zip_map = {}
    try:
        path = resource_path("Zipcodes-with-Reps-Complete.xlsx")
        if not os.path.exists(path):
            print(f"WARNING: Lookup file not found at {path}")
            return zip_map
        df = pd.read_excel(path, sheet_name="Full Scrape", dtype=str)
        df.columns = [c.strip() for c in df.columns]
        for _, row in df.iterrows():
            raw_zip = str(row.get("Zipcode", "")).strip()
            zipcode = raw_zip.split("-")[0].strip().zfill(5)
            if len(zipcode) != 5 or not zipcode.isdigit():
                continue
            council  = str(row.get("City Council District",   "")).strip()
            assembly = str(row.get("State Assembly District", "")).strip()
            senate   = str(row.get("State Senate District",   "")).strip()
            congress = str(row.get("US House District",       "")).strip()
            if not council or council.lower() == "nan":
                continue
            entry = {
                "City Council District":   council,
                "State Assembly District": assembly,
                "State Senate District":   senate,
                "US House District":       congress,
            }
            if zipcode not in zip_map:
                zip_map[zipcode] = []
            if entry not in zip_map[zipcode]:
                zip_map[zipcode].append(entry)
        print(f"ZIP lookup loaded: {len(zip_map)} unique zip codes.")
    except Exception as e:
        print(f"Error loading ZIP lookup: {e}")
    return zip_map


# ─────────────────────────────────────────────────────────────
#  Scrape mygovnyc.org
# ─────────────────────────────────────────────────────────────
def scrape_mygovnyc(address, city, state="NY"):
    result = {
        "City Council District": "", "State Assembly District": "",
        "State Senate District": "", "US House District": ""
    }
    try:
        full_addr = f"{address}, {city}, {state}"
        headers = {
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/120.0.0.0 Safari/537.36"
            ),
            "Accept": "text/html,application/xhtml+xml,*/*;q=0.8",
            "Accept-Language": "en-US,en;q=0.9",
        }
        resp = requests.get(
            "https://www.mygovnyc.org/",
            params={"address": full_addr},
            headers=headers,
            timeout=12
        )
        if resp.status_code != 200:
            return result
        soup = BeautifulSoup(resp.text, "lxml")
        for el in soup.find_all(string=re.compile(r"city council", re.I)):
            text = el.parent.get_text(" ", strip=True)
            m = re.search(r"district\s+(\d+)", text, re.I)
            if m:
                result["City Council District"] = f"District {m.group(1)}"
                break
        for el in soup.find_all(string=re.compile(r"assembly", re.I)):
            text = el.parent.get_text(" ", strip=True)
            m = re.search(r"district\s+(\d+)", text, re.I)
            if m:
                result["State Assembly District"] = f"Assembly District {m.group(1)}"
                break
        for el in soup.find_all(string=re.compile(r"state senate", re.I)):
            text = el.parent.get_text(" ", strip=True)
            m = re.search(r"district\s+(\d+)", text, re.I)
            if m:
                result["State Senate District"] = f"Senate District {m.group(1)}"
                break
        for el in soup.find_all(string=re.compile(r"congressional|us house|u\.s\. house", re.I)):
            text = el.parent.get_text(" ", strip=True)
            m = re.search(r"(NY\s*\d+|\d+(?:st|nd|rd|th)\s+district)", text, re.I)
            if m:
                result["US House District"] = m.group(1).strip()
                break
    except Exception as e:
        print(f"Scrape error for '{address}': {e}")
    return result


# ─────────────────────────────────────────────────────────────
#  Scrape representative names from official sites
# ─────────────────────────────────────────────────────────────
def scrape_rep_names():
    names = {"council": {}, "assembly": {}, "senate": {}, "congress": {}}
    _hdr  = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"}

    try:
        r = requests.get("https://council.nyc.gov/members/", headers=_hdr, timeout=15)
        soup = BeautifulSoup(r.text, "lxml")
        for card in soup.find_all(class_=re.compile(r"council-member|member-card", re.I)):
            txt = card.get_text(" ", strip=True)
            m_d = re.search(r"district\s+(\d+)", txt, re.I)
            m_n = re.search(r"(?:Council Member|CM)?\s*([A-Z][a-z]+(?: [A-Z][a-z'-]+)+)", txt)
            if m_d and m_n:
                names["council"][m_d.group(1)] = m_n.group(1).strip()
        if not names["council"]:
            for el in soup.find_all(string=re.compile(r"District\s+\d+", re.I)):
                parent = el.parent
                m_d = re.search(r"District\s+(\d+)", el, re.I)
                if m_d:
                    sibling = parent.find_next(["h2", "h3", "p", "a"])
                    if sibling:
                        candidate = sibling.get_text(strip=True)
                        if re.match(r"[A-Z][a-z]", candidate):
                            names["council"][m_d.group(1)] = candidate
    except Exception as e:
        print(f"Council name scrape: {e}")

    try:
        r    = requests.get("https://www.nyassembly.gov/mem/", headers=_hdr, timeout=15)
        soup = BeautifulSoup(r.text, "lxml")
        for row in soup.find_all("tr"):
            cells = row.find_all(["td", "th"])
            if len(cells) >= 2:
                d_txt = cells[0].get_text(strip=True)
                n_txt = cells[1].get_text(strip=True)
                m = re.search(r"(\d+)", d_txt)
                if m and re.match(r"[A-Z]", n_txt):
                    names["assembly"][m.group(1)] = n_txt
    except Exception as e:
        print(f"Assembly name scrape: {e}")

    try:
        r    = requests.get("https://www.nysenate.gov/senators-committees",
                            headers=_hdr, timeout=15)
        soup = BeautifulSoup(r.text, "lxml")
        for card in soup.find_all(class_=re.compile(r"senator|nys-senator|c-senator", re.I)):
            txt = card.get_text(" ", strip=True)
            m_d = re.search(r"(?:District|SD)\s*[:\-]?\s*(\d+)", txt, re.I)
            m_n = re.search(r"Senator\s+([A-Z][a-z]+(?: [A-Z][a-z'-]+)+)", txt)
            if not m_n:
                m_n = re.search(r"([A-Z][a-z]+(?: [A-Z][a-z'-]+)+)", txt)
            if m_d and m_n:
                names["senate"][m_d.group(1)] = m_n.group(1).strip()
    except Exception as e:
        print(f"Senate name scrape: {e}")

    try:
        r    = requests.get("https://www.house.gov/representatives", headers=_hdr, timeout=15)
        soup = BeautifulSoup(r.text, "lxml")
        ny_header = None
        for tag in soup.find_all(["h2", "h3", "h4", "caption"]):
            if "new york" in tag.get_text(strip=True).lower():
                ny_header = tag
                break
        if ny_header:
            tbl = ny_header.find_next("table")
            if tbl:
                for tr in tbl.find_all("tr")[1:]:
                    cells = tr.find_all("td")
                    if len(cells) >= 2:
                        d_raw = cells[0].get_text(strip=True)
                        n_raw = cells[1].get_text(strip=True)
                        m = re.search(r"(\d+)", d_raw)
                        if m:
                            names["congress"][f"NY{m.group(1)}"] = n_raw
    except Exception as e:
        print(f"Congress name scrape: {e}")

    print(f"Rep names loaded — council:{len(names['council'])} "
          f"assembly:{len(names['assembly'])} "
          f"senate:{len(names['senate'])} "
          f"congress:{len(names['congress'])}")
    return names


# ─────────────────────────────────────────────────────────────
#  Background threads
# ─────────────────────────────────────────────────────────────
class RepNameWorker(QThread):
    names_ready = pyqtSignal(object)
    def run(self):
        try:
            self.names_ready.emit(scrape_rep_names())
        except Exception as e:
            print(f"RepNameWorker: {e}")
            self.names_ready.emit({"council": {}, "assembly": {}, "senate": {}, "congress": {}})


class LookupWorker(QThread):
    progress       = pyqtSignal(int, str)
    result_ready   = pyqtSignal(object)
    error_occurred = pyqtSignal(str)

    def __init__(self, df, zip_lookup):
        super().__init__()
        self.df         = df.copy()
        self.zip_lookup = zip_lookup
        self._cancelled = False

    def cancel(self):
        self._cancelled = True

    def run(self):
        try:
            cols  = list(self.df.columns)
            total = len(self.df)

            zip_col   = detect_col(cols, "zip")
            addr_col  = detect_col(cols, "address")
            city_col  = detect_col(cols, "city")
            state_col = detect_col(cols, "state")

            council_list  = []
            assembly_list = []
            senate_list   = []
            congress_list = []
            method_list   = []

            for i, (_, row) in enumerate(self.df.iterrows()):
                if self._cancelled:
                    break

                raw_zip = str(row[zip_col]).strip()  if zip_col  else ""
                address = str(row[addr_col]).strip() if addr_col else ""
                city    = str(row[city_col]).strip() if city_col else ""
                state   = str(row[state_col]).strip().upper() if state_col else "NY"

                address = re.sub(r"^['\-\s]+", "", address).strip()
                if address.lower() == "nan":
                    address = ""

                zipcode = raw_zip.split("-")[0].strip()
                if zipcode and len(zipcode) < 5:
                    zipcode = zipcode.zfill(5)

                council = assembly = senate = congress = ""
                method  = "none"

                # PRIMARY: full address scrape (NY only, valid addresses)
                if is_valid_address(address) and state in ("NY", "NEW YORK"):
                    self.progress.emit(
                        int(i / total * 100),
                        f"[Address] {i+1}/{total}: {address[:45]}..."
                    )
                    scraped  = scrape_mygovnyc(address, city, state)
                    council  = scraped["City Council District"]
                    assembly = scraped["State Assembly District"]
                    senate   = scraped["State Senate District"]
                    congress = scraped["US House District"]
                    if council:
                        method = "address"
                    time.sleep(0.4)

                # FALLBACK: ZIP lookup if address failed or invalid
                if not council and zipcode in self.zip_lookup:
                    matches = self.zip_lookup[zipcode]
                    if matches:
                        d        = matches[0]
                        council  = d.get("City Council District",   "")
                        assembly = d.get("State Assembly District", "")
                        senate   = d.get("State Senate District",   "")
                        congress = d.get("US House District",       "")
                        method   = "zip_multi" if len(matches) > 1 else "zip"

                council_list.append(council)
                assembly_list.append(assembly)
                senate_list.append(senate)
                congress_list.append(congress)
                method_list.append(method)

                self.progress.emit(
                    int((i + 1) / total * 100),
                    f"Processing {i+1:,} of {total:,}   ZIP:{zipcode}  [{method}]"
                )

            result = self.df.copy()
            result["City Council District"]   = council_list
            result["State Assembly District"] = assembly_list
            result["State Senate District"]   = senate_list
            result["US House District"]       = congress_list
            result["_match_method"]           = method_list

            self.result_ready.emit(result)

        except Exception as e:
            import traceback
            self.error_occurred.emit(traceback.format_exc())


# ─────────────────────────────────────────────────────────────
#  Stat card widget
# ─────────────────────────────────────────────────────────────
class StatCard(QFrame):
    def __init__(self, label, value="—", accent="#1565C0"):
        super().__init__()
        self.setObjectName("statCard")
        lay = QVBoxLayout(self)
        lay.setContentsMargins(16, 12, 16, 12)
        lay.setSpacing(4)

        self.val_lbl = QLabel(value)
        self.val_lbl.setAlignment(Qt.AlignCenter)
        self.val_lbl.setFont(QFont("Segoe UI", 22, QFont.Bold))
        self.val_lbl.setStyleSheet(f"color:{accent};background:transparent;")

        self.key_lbl = QLabel(label)
        self.key_lbl.setAlignment(Qt.AlignCenter)
        self.key_lbl.setStyleSheet(
            "color:#546E7A;font-size:11px;font-weight:600;"
            "letter-spacing:0.5px;background:transparent;"
        )
        lay.addWidget(self.val_lbl)
        lay.addWidget(self.key_lbl)

    def set_value(self, v):
        self.val_lbl.setText(v)


# ─────────────────────────────────────────────────────────────
#  Main Window
# ─────────────────────────────────────────────────────────────
class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.df          = None
        self.result_df   = None
        self.worker      = None
        self.name_worker = None
        self.rep_names   = {"council": {}, "assembly": {}, "senate": {}, "congress": {}}
        self.zip_lookup  = load_zip_lookup()
        self._setup_ui()
        self._apply_theme()
        self._start_rep_name_fetch()

    def _start_rep_name_fetch(self):
        self.status_lbl.setText("⏳  Loading representative names in background…")
        self.name_worker = RepNameWorker()
        self.name_worker.names_ready.connect(self._on_names_ready)
        self.name_worker.start()

    def _on_names_ready(self, names):
        self.rep_names = names
        total = sum(len(v) for v in names.values())
        self.status_lbl.setText(
            f"✅  Ready  —  {total} representative names loaded.  "
            f"Load a CSV or Excel file to begin."
        )

    # ── UI Construction ──────────────────────────────────────
    def _setup_ui(self):
        self.setWindowTitle("NYC RepTracker  —  QEDC")
        self.setMinimumSize(1200, 800)

        ico = resource_path("icon_256x256.ico")
        if os.path.exists(ico):
            self.setWindowIcon(QIcon(ico))

        root   = QWidget()
        self.setCentralWidget(root)
        master = QVBoxLayout(root)
        master.setContentsMargins(0, 0, 0, 0)
        master.setSpacing(0)

        # ── Nav bar (full width, no cutouts) ─────────────────
        nav = QFrame()
        nav.setObjectName("navBar")
        nav.setFixedHeight(90)
        nav.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        nav_lay = QHBoxLayout(nav)
        nav_lay.setContentsMargins(28, 0, 28, 0)
        nav_lay.setSpacing(0)

        # NYC RepTracker generated title image
        title_lbl = QLabel()
        title_lbl.setStyleSheet("background: transparent;")
        title_img_path = resource_path("generated-image.png")
        if os.path.exists(title_img_path):
            px = QPixmap(title_img_path).scaled(
                420, 72, Qt.KeepAspectRatio, Qt.SmoothTransformation
            )
            title_lbl.setPixmap(px)
        else:
            title_lbl.setText("NYC RepTracker")
            title_lbl.setStyleSheet(
                "font-size:26px;font-weight:700;color:#1565C0;background:transparent;"
            )

        # Support button
        self.support_btn = QPushButton("💬  Support")
        self.support_btn.setObjectName("btnSupport")
        self.support_btn.setFixedHeight(40)
        self.support_btn.setToolTip(
            "Victor Prado — vprado@queensny.org\nClick to copy email"
        )
        self.support_btn.clicked.connect(self._copy_support_email)

        nav_lay.addWidget(title_lbl)
        nav_lay.addStretch()
        nav_lay.addWidget(self.support_btn)

        # ── Content area ─────────────────────────────────────
        content_wrap = QWidget()
        content_wrap.setObjectName("contentArea")
        content = QVBoxLayout(content_wrap)
        content.setContentsMargins(28, 20, 28, 12)
        content.setSpacing(14)

        # Toolbar card
        toolbar = QFrame()
        toolbar.setObjectName("toolCard")
        tb_lay  = QHBoxLayout(toolbar)
        tb_lay.setContentsMargins(20, 14, 20, 14)
        tb_lay.setSpacing(12)

        self.load_btn = QPushButton("  📂  Load File")
        self.load_btn.setObjectName("btnPrimary")
        self.load_btn.setFixedHeight(42)
        self.load_btn.clicked.connect(self.load_file)

        self.file_lbl = QLabel("No file loaded — supported formats: CSV, XLSX")
        self.file_lbl.setObjectName("fileLabel")
        self.file_lbl.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)

        self.run_btn = QPushButton("  ▶  Run Lookup")
        self.run_btn.setObjectName("btnSuccess")
        self.run_btn.setFixedHeight(42)
        self.run_btn.setEnabled(False)
        self.run_btn.clicked.connect(self.run_lookup)

        self.export_btn = QPushButton("  💾  Export Excel")
        self.export_btn.setObjectName("btnWarning")
        self.export_btn.setFixedHeight(42)
        self.export_btn.setEnabled(False)
        self.export_btn.clicked.connect(self.export_results)

        for w in [self.load_btn, self.file_lbl, self.run_btn, self.export_btn]:
            tb_lay.addWidget(w)

        # Progress card
        prog_card = QFrame()
        prog_card.setObjectName("toolCard")
        prog_lay  = QVBoxLayout(prog_card)
        prog_lay.setContentsMargins(20, 10, 20, 10)
        prog_lay.setSpacing(5)

        self.status_lbl = QLabel("Initializing…")
        self.status_lbl.setObjectName("statusLabel")

        self.prog_bar = QProgressBar()
        self.prog_bar.setFixedHeight(8)
        self.prog_bar.setTextVisible(False)
        self.prog_bar.setVisible(False)

        prog_lay.addWidget(self.status_lbl)
        prog_lay.addWidget(self.prog_bar)

        # Stat cards row
        stats_row = QHBoxLayout()
        stats_row.setSpacing(12)
        self.card_total   = StatCard("TOTAL RECORDS",    "—", "#1565C0")
        self.card_matched = StatCard("MATCHED",          "—", "#2E7D32")
        self.card_zip_fb  = StatCard("ZIP FALLBACK",     "—", "#E65100")
        self.card_cache   = StatCard("ZIP CACHE ENTRIES",
                                     f"{len(self.zip_lookup):,}", "#6A1B9A")
        for c in [self.card_total, self.card_matched,
                  self.card_zip_fb, self.card_cache]:
            stats_row.addWidget(c)

        # Section header
        tbl_hdr = QLabel("Results Preview")
        tbl_hdr.setObjectName("sectionHeader")

        # Results table
        self.table = QTableWidget()
        self.table.setObjectName("dataTable")
        self.table.horizontalHeader().setSectionResizeMode(
            QHeaderView.ResizeToContents)
        self.table.setAlternatingRowColors(True)
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.setShowGrid(False)
        self.table.verticalHeader().setDefaultSectionSize(32)
        self.table.verticalHeader().setVisible(False)

        # ── Footer — QEDC logo bottom-left ───────────────────
        footer_widget = QWidget()
        footer_widget.setObjectName("footerWidget")
        footer_lay = QHBoxLayout(footer_widget)
        footer_lay.setContentsMargins(4, 4, 4, 4)
        footer_lay.setSpacing(12)

        qedc_foot_lbl = QLabel()
        qedc_foot_lbl.setStyleSheet("background: transparent;")
        qedc_logo_path = resource_path("QEDC - Full Logo (Primary Color).png")
        if os.path.exists(qedc_logo_path):
            px_f = QPixmap(qedc_logo_path).scaled(
                130, 44, Qt.KeepAspectRatio, Qt.SmoothTransformation
            )
            qedc_foot_lbl.setPixmap(px_f)
        else:
            qedc_foot_lbl.setText("QEDC")
            qedc_foot_lbl.setStyleSheet(
                "font-size:14px;font-weight:bold;color:#1565C0;background:transparent;"
            )

        footer_text = QLabel(
            "Queens Economic Development Corporation  •  NYC RepTracker  "
            "•  Support: Victor Prado — vprado@queensny.org"
        )
        footer_text.setObjectName("footerLabel")
        footer_text.setAlignment(Qt.AlignCenter)
        footer_text.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)

        footer_spacer = QLabel()
        footer_spacer.setFixedWidth(130)
        footer_spacer.setStyleSheet("background: transparent;")

        footer_lay.addWidget(qedc_foot_lbl)
        footer_lay.addWidget(footer_text)
        footer_lay.addWidget(footer_spacer)

        # Assemble content
        content.addWidget(toolbar)
        content.addWidget(prog_card)
        content.addLayout(stats_row)
        content.addWidget(tbl_hdr)
        content.addWidget(self.table)
        content.addWidget(footer_widget)
        content.setStretch(4, 1)

        master.addWidget(nav)
        master.addWidget(content_wrap)

    # ── Theme ────────────────────────────────────────────────
    def _apply_theme(self):
        self.setStyleSheet("""
        QMainWindow, QWidget {
            background: #F4F6F9;
            color: #263238;
            font-family: 'Segoe UI', 'Helvetica Neue', Arial, sans-serif;
            font-size: 13px;
        }
        #navBar {
            background: #F4F6F9;
            border-bottom: 2px solid #DDE3ED;
        }
        #navBar QLabel  { background: transparent; }
        #navBar QWidget { background: transparent; }
        #navBar QPushButton { background: #6A1B9A; }

        #contentArea { background: #F4F6F9; }

        #toolCard {
            background: #FFFFFF;
            border-radius: 10px;
            border: 1px solid #DDE3ED;
        }
        #statCard {
            background: #FFFFFF;
            border-radius: 10px;
            border: 1px solid #DDE3ED;
            min-width: 160px;
        }

        QPushButton {
            padding: 0 20px;
            border-radius: 7px;
            font-weight: 600;
            font-size: 13px;
            border: none;
            min-width: 130px;
        }
        #btnPrimary           { background: #1565C0; color: #FFFFFF; }
        #btnPrimary:hover     { background: #1976D2; }
        #btnPrimary:pressed   { background: #0D47A1; }

        #btnSuccess           { background: #2E7D32; color: #FFFFFF; }
        #btnSuccess:hover     { background: #388E3C; }
        #btnSuccess:pressed   { background: #1B5E20; }
        #btnSuccess:disabled  { background: #BDBDBD; color: #757575; }

        #btnWarning           { background: #E65100; color: #FFFFFF; }
        #btnWarning:hover     { background: #F4511E; }
        #btnWarning:pressed   { background: #BF360C; }
        #btnWarning:disabled  { background: #BDBDBD; color: #757575; }

        #btnSupport           { background: #6A1B9A; color: #FFFFFF; min-width: 110px; }
        #btnSupport:hover     { background: #7B1FA2; }
        #btnSupport:pressed   { background: #4A148C; }

        #fileLabel  { color: #455A64; font-size: 12px; padding: 0 8px; }

        QProgressBar {
            background: #E3E8EF;
            border-radius: 4px;
            border: none;
        }
        QProgressBar::chunk {
            background: qlineargradient(
                x1:0, y1:0, x2:1, y2:0,
                stop:0 #1565C0, stop:1 #42A5F5
            );
            border-radius: 4px;
        }

        #statusLabel   { color: #546E7A; font-size: 12px; }
        #sectionHeader { font-size: 14px; font-weight: 700; color: #37474F; padding: 4px 2px; }

        #dataTable {
            background: #FFFFFF;
            alternate-background-color: #F8FAFC;
            border: 1px solid #DDE3ED;
            border-radius: 10px;
            gridline-color: transparent;
            outline: none;
        }
        QHeaderView::section {
            background: #1565C0;
            color: #FFFFFF;
            padding: 10px 12px;
            border: none;
            font-weight: 700;
            font-size: 12px;
            letter-spacing: 0.4px;
        }
        QHeaderView::section:first { border-top-left-radius: 8px; }
        QHeaderView::section:last  { border-top-right-radius: 8px; }
        QTableWidget::item {
            padding: 6px 10px;
            border-bottom: 1px solid #EEF1F5;
        }
        QTableWidget::item:selected { background: #BBDEFB; color: #0D47A1; }

        QScrollBar:vertical {
            background: #F4F6F9; width: 8px; border-radius: 4px;
        }
        QScrollBar::handle:vertical {
            background: #B0BEC5; border-radius: 4px; min-height: 30px;
        }
        QScrollBar::handle:vertical:hover { background: #78909C; }
        QScrollBar::add-line:vertical,
        QScrollBar::sub-line:vertical { height: 0; }
        QScrollBar:horizontal {
            background: #F4F6F9; height: 8px; border-radius: 4px;
        }
        QScrollBar::handle:horizontal { background: #B0BEC5; border-radius: 4px; }
        QScrollBar::add-line:horizontal,
        QScrollBar::sub-line:horizontal { width: 0; }

        #footerWidget { background: #F4F6F9; }
        #footerLabel  { color: #90A4AE; font-size: 11px; padding: 4px 0; }

        QMessageBox { background: #FFFFFF; }
        QMessageBox QPushButton {
            min-width: 80px; padding: 6px 16px;
            background: #1565C0; color: white; border-radius: 6px;
        }
        QToolTip {
            background: #FFFFFF; color: #263238;
            border: 1px solid #DDE3ED; border-radius: 6px;
            padding: 6px 10px; font-size: 12px;
        }
        """)

    # ── Support button ───────────────────────────────────────
    def _copy_support_email(self):
        email = "vprado@queensny.org"
        QApplication.clipboard().setText(email)
        QMessageBox.information(
            self,
            "Support Contact",
            f"<b>Victor Prado</b><br>"
            f"Queens Economic Development Corporation<br><br>"
            f"📧  <b>{email}</b><br><br>"
            f"✅  Email address copied to clipboard!"
        )

    # ── Load file ────────────────────────────────────────────
    def load_file(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Open Data File", "",
            "Data Files (*.csv *.xlsx *.xls);;CSV (*.csv);;Excel (*.xlsx *.xls)"
        )
        if not path:
            return
        try:
            if path.lower().endswith((".xlsx", ".xls")):
                self.df = pd.read_excel(path, dtype=str)
            else:
                self.df = pd.read_csv(path, dtype=str, encoding="utf-8-sig")
            self.df.fillna("", inplace=True)
            self._show_table(self.df.head(100))
            self.file_lbl.setText(
                f"📄  {os.path.basename(path)}   ({len(self.df):,} rows)"
            )
            self.run_btn.setEnabled(True)
            self.export_btn.setEnabled(False)
            self.result_df = None
            self.card_total.set_value(f"{len(self.df):,}")
            self.card_matched.set_value("—")
            self.card_zip_fb.set_value("—")
            self.status_lbl.setText(
                f"✅  File loaded — {len(self.df):,} rows.  "
                f"Click  ▶ Run Lookup  to begin."
            )
        except Exception as e:
            QMessageBox.critical(self, "Load Error", str(e))

    # ── Table display ────────────────────────────────────────
    def _show_table(self, df):
        self.table.clear()
        if df is None or df.empty:
            return
        HIGHLIGHT = {
            "City Council District", "State Assembly District",
            "State Senate District", "US House District"
        }
        SKIP = {"_match_method"}
        cols = [c for c in df.columns if c not in SKIP]
        self.table.setRowCount(len(df))
        self.table.setColumnCount(len(cols))
        self.table.setHorizontalHeaderLabels(cols)
        for r in range(len(df)):
            for c, col_name in enumerate(cols):
                val = str(df.iloc[r][col_name])
                if val in ("nan", ""):
                    val = ""
                item = QTableWidgetItem(val)
                if col_name in HIGHLIGHT:
                    if val:
                        item.setForeground(QColor("#1B5E20"))
                        item.setBackground(QColor("#F1F8E9"))
                    else:
                        item.setForeground(QColor("#B71C1C"))
                        item.setText("—")
                        item.setBackground(QColor("#FFF3E0"))
                self.table.setItem(r, c, item)

    # ── Run lookup ───────────────────────────────────────────
    def run_lookup(self):
        if self.df is None:
            return
        self.run_btn.setEnabled(False)
        self.export_btn.setEnabled(False)
        self.prog_bar.setVisible(True)
        self.prog_bar.setValue(0)
        self.status_lbl.setText("Starting lookup…")
        self.worker = LookupWorker(self.df, self.zip_lookup)
        self.worker.progress.connect(self._on_progress)
        self.worker.result_ready.connect(self._on_result)
        self.worker.error_occurred.connect(self._on_error)
        self.worker.start()

    def _on_progress(self, pct, msg):
        self.prog_bar.setValue(pct)
        self.status_lbl.setText(msg)

    def _on_result(self, df):
        self.result_df = df
        d_col      = "City Council District"
        method_col = "_match_method"
        matched    = (df[d_col].str.strip() != "").sum() if d_col in df.columns else 0
        zip_fb     = (df[method_col].isin(["zip", "zip_multi"])).sum() \
                       if method_col in df.columns else 0
        self._show_table(df.head(200))
        self.card_total.set_value(f"{len(df):,}")
        self.card_matched.set_value(f"{matched:,}")
        self.card_zip_fb.set_value(f"{zip_fb:,}")
        self.prog_bar.setValue(100)
        self.status_lbl.setText(
            f"✅  Lookup complete  —  {matched:,} matched  |  "
            f"{zip_fb:,} via ZIP fallback  |  "
            f"{len(df)-matched:,} unmatched  (showing first 200 rows)"
        )
        self.run_btn.setEnabled(True)
        self.export_btn.setEnabled(True)

    def _on_error(self, msg):
        QMessageBox.critical(self, "Lookup Error", msg)
        self.run_btn.setEnabled(True)
        self.prog_bar.setVisible(False)

    # ── Export with Summary sheet ────────────────────────────
    def export_results(self):
        if self.result_df is None:
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "Save Results", "NYC_RepTracker_Results.xlsx",
            "Excel Workbook (*.xlsx)"
        )
        if not path:
            return
        try:
            export_df  = self.result_df.drop(columns=["_match_method"], errors="ignore")
            summary_df = self._build_summary(self.result_df)
            with pd.ExcelWriter(path, engine="openpyxl") as writer:
                export_df.to_excel(writer, index=False, sheet_name="Results")
                summary_df.to_excel(writer, index=False, sheet_name="Summary")
            self._format_excel(path)
            QMessageBox.information(self, "Export Successful", f"File saved:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Save Error", str(e))

    # ── Build Summary DataFrame ──────────────────────────────
    def _build_summary(self, df):
        zip_col = detect_col(list(df.columns), "zip")
        council_counts  = {}
        assembly_counts = {}
        senate_counts   = {}
        congress_counts = {}

        for _, row in df.iterrows():
            method   = str(row.get("_match_method", "address"))
            council  = str(row.get("City Council District",   "")).strip()
            assembly = str(row.get("State Assembly District", "")).strip()
            senate   = str(row.get("State Senate District",   "")).strip()
            congress = str(row.get("US House District",       "")).strip()

            if method == "zip_multi" and zip_col:
                raw_zip = str(row[zip_col]).strip()
                zipcode = raw_zip.split("-")[0].strip().zfill(5)
                for match in self.zip_lookup.get(zipcode, []):
                    c = match.get("City Council District",   "")
                    a = match.get("State Assembly District", "")
                    s = match.get("State Senate District",   "")
                    g = match.get("US House District",       "")
                    if c: council_counts[c]  = council_counts.get(c, 0)  + 1
                    if a: assembly_counts[a] = assembly_counts.get(a, 0) + 1
                    if s: senate_counts[s]   = senate_counts.get(s, 0)   + 1
                    if g: congress_counts[g] = congress_counts.get(g, 0) + 1
            else:
                if council:  council_counts[council]   = council_counts.get(council, 0)   + 1
                if assembly: assembly_counts[assembly] = assembly_counts.get(assembly, 0) + 1
                if senate:   senate_counts[senate]     = senate_counts.get(senate, 0)     + 1
                if congress: congress_counts[congress] = congress_counts.get(congress, 0) + 1

        def sort_key(d_str):
            m = re.search(r"(\d+)", d_str)
            return int(m.group(1)) if m else 9999

        def get_name(dist, category):
            m = re.search(r"(\d+)", dist)
            if not m:
                return "—"
            key = m.group(1)
            if category == "congress":
                key = dist.upper().replace(" ", "")
            return self.rep_names.get(category, {}).get(key, "—")

        rows = []
        for section, counts, cat in [
            ("NYC City Council",  council_counts,  "council"),
            ("NY State Assembly", assembly_counts, "assembly"),
            ("NY State Senate",   senate_counts,   "senate"),
            ("US House (NY)",     congress_counts, "congress"),
        ]:
            rows.append({
                "Section": section, "District": "",
                "Representative Name": "", "Record Count": "",
            })
            for dist, cnt in sorted(counts.items(), key=lambda x: sort_key(x[0])):
                rows.append({
                    "Section":            "",
                    "District":           dist,
                    "Representative Name": get_name(dist, cat),
                    "Record Count":       cnt,
                })
            rows.append({
                "Section": "", "District": "",
                "Representative Name": "", "Record Count": "",
            })

        return pd.DataFrame(rows)

    # ── Format Excel workbook ────────────────────────────────
    def _format_excel(self, path):
        wb    = openpyxl.load_workbook(path)
        blue  = PatternFill("solid", fgColor="1565C0")
        green = PatternFill("solid", fgColor="2E7D32")
        hdr_font   = XLFont(bold=True, color="FFFFFF", size=11)
        sec_font   = XLFont(bold=True, color="FFFFFF", size=11)
        center     = XLAlign(horizontal="center", vertical="center")
        left       = XLAlign(horizontal="left",   vertical="center")

        ws = wb["Results"]
        for cell in ws[1]:
            cell.fill      = blue
            cell.font      = hdr_font
            cell.alignment = center
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)
        ws.freeze_panes = "A2"

        ws2 = wb["Summary"]
        for cell in ws2[1]:
            cell.fill      = blue
            cell.font      = hdr_font
            cell.alignment = center
        for row in ws2.iter_rows(min_row=2):
            section_val = ws2.cell(row=row[0].row, column=1).value
            for cell in row:
                if section_val and str(section_val).strip():
                    cell.fill      = green
                    cell.font      = sec_font
                    cell.alignment = left
                else:
                    cell.alignment = center if cell.column == 4 else left
        for col in ws2.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws2.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 60)
        ws2.freeze_panes = "A2"

        wb.save(path)


# ─────────────────────────────────────────────────────────────
def main():
    app = QApplication(sys.argv)
    app.setStyle("Fusion")
    win = MainWindow()
    win.show()
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()
